from openpyxl import Workbook
from datetime import datetime
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import Pairs, Trade, Order, Deal, FutureTrades
from django.db.models import Q
from account.oauth import verify_token


class Mt5FileViewSet(viewsets.ViewSet):
    def trades_excel(self,request):
        """
        return user trades/orders/deals in excel
        """
        try:
            token = request.data.pop('token')
        except:
            return Response({'message': 'invalid data'}, status=status.HTTP_403_FORBIDDEN)
        
        payload =  verify_token(token)
        if not payload:
            return Response({'message': 'error verifying token'}, status=status.HTTP_403_FORBIDDEN)

        trades = Trade.objects.select_related('user').filter(user__id=payload['id'], state='close')        

        filters = Q()
        if 'max_date' in request.data.keys():
            max_date =['max_date']
            max_date = datetime.strptime(max_date, '%Y-%m-%d').date()
            filters &= Q(time__date__lte=max_date)
        
        if 'min_date' in request.data.keys():
            min_date = request.data['max_date']
            min_date = datetime.strptime(min_date, '%Y-%m-%d').date()
            filters |= Q(time__date__gte=min_date)
        
        trades = trades.filter(filters)

        workbook = Workbook()
        sheet = workbook.active
        header = ["ticket", "symbol", "type", "entry", "exit", "profit", "unit", "tp", "sl", "leverage", "date", "commission", "swap"]
        sheet.append(header)

        for trade in trades:
            trade_time=trade.time.strftime('%Y-%m-%d %H:%M:%S') if trade.time else ''
            sheet.append( [trade.ticket, trade.symbol.name, trade.type, trade.entry, trade.exit, trade.profit, trade.unit, trade.tp, trade.sl, trade.leverage, trade_time, trade.commission, trade.swap] )
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        workbook.save(response)
        
        return response
    
    def deals_excel(self,request):
        """
        return user trades/orders/deals in excel
        """
        try:
            token = request.data.pop('token')
        except:
            return Response({'message': 'invalid data'}, status=status.HTTP_403_FORBIDDEN)
        
        payload =  verify_token(token)
        if not payload:
            return Response({'message': 'error verifying token'}, status=status.HTTP_403_FORBIDDEN)

        deals = Deal.objects.select_related('user').filter(user__id=payload['id'])
        
        filters = Q()
        if 'max_date' in request.data.keys():
            max_date =['max_date']
            max_date = datetime.strptime(max_date, '%Y-%m-%d').date()
            filters &= Q(time__date__lte=max_date)
        
        if 'min_date' in request.data.keys():
            min_date = request.data['max_date']
            min_date = datetime.strptime(min_date, '%Y-%m-%d').date()
            filters |= Q(time__date__gte=min_date)
        
        deals = deals.filter(filters)
        
        workbook = Workbook()
        sheet = workbook.active
        header = ["ticket", "symbol", "trade_id", "type", "dir", "price", "profit", "unit", "commission", "date"]
        sheet.append(header)
        for deal in deals:
            deal_time=deal.time.strftime('%Y-%m-%d %H:%M:%S') if deal.time else ''
            sheet.append( [deal.ticket, deal.symbol.name, deal.trade_id, deal.type, deal.dir, deal.price, deal.profit, deal.unit, deal.commission, deal_time] )
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        workbook.save(response)
        
        return response
    
    def orders_excel(self,request):
        """
        return user trades/orders/deals in excel
        """
        try:
            token = request.data.pop('token')
        except:
            return Response({'message': 'invalid data'}, status=status.HTTP_403_FORBIDDEN)
        
        payload =  verify_token(token)
        if not payload:
            return Response({'message': 'error verifying token'}, status=status.HTTP_403_FORBIDDEN)

        orders = Order.objects.select_related('user').filter(user__id=payload['id'])
        
        filters = Q()
        if 'max_date' in request.data.keys():
            max_date =['max_date']
            max_date = datetime.strptime(max_date, '%Y-%m-%d').date()
            filters &= Q(time__date__lte=max_date)
        
        if 'min_date' in request.data.keys():
            min_date = request.data['max_date']
            min_date = datetime.strptime(min_date, '%Y-%m-%d').date()
            filters |= Q(time__date__gte=min_date)
        
        orders = orders.filter(filters)

        workbook = Workbook()
        sheet = workbook.active
        header = ["ticket", "symbol", "type", "price", "unit", "date", "result"]
        sheet.append(header)
        for order in orders:
            order_time=order.time.strftime('%Y-%m-%d %H:%M:%S') if order.time else ''
            sheet.append( [order.ticket, order.symbol.name, order.type, order.price, order.unit, order_time, order.result] )
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        workbook.save(response)
        
        return response

